# Path: xbrl_parser/output/excel_exporter.py
"""
Excel Export Utilities

Export XBRL data to Excel workbooks with multiple sheets and formatting.

Requires openpyxl package for Excel file creation.
"""

import logging
from pathlib import Path
from datetime import datetime

from xbrl_parser.models.parsed_filing import ParsedFiling
from output.extracted_data.data_extractor import DataExtractor


class ExcelExporter:
    """
    Export XBRL filings to Excel workbooks.

    Creates Excel files with multiple sheets:
    - Summary: Filing metadata and statistics
    - Facts: All facts with context details
    - Contexts: Context definitions
    - Units: Unit definitions
    - Errors: Parsing errors and warnings

    Example:
        exporter = ExcelExporter()
        exporter.export(filing, 'output.xlsx')
    """

    # Column width constants
    COLUMN_WIDTH_NARROW: int = 15
    COLUMN_WIDTH_MEDIUM: int = 20
    COLUMN_WIDTH_WIDE: int = 30
    COLUMN_WIDTH_EXTRA_WIDE: int = 40
    COLUMN_WIDTH_VERY_WIDE: int = 50

    def __init__(self):
        """Initialize Excel exporter."""
        self.logger = logging.getLogger(__name__)
        self.extractor = DataExtractor()
        
        # Check if openpyxl available
        try:
            import openpyxl
            self.openpyxl = openpyxl
            self.has_openpyxl = True
        except ImportError:
            self.openpyxl = None
            self.has_openpyxl = False
            self.logger.warning(
                "openpyxl not available. Excel export disabled. "
                "Install with: pip install openpyxl"
            )
    
    def export(
        self,
        filing: ParsedFiling,
        output_path: Path,
        include_errors: bool = True
    ) -> None:
        """
        Export filing to Excel workbook.
        
        Args:
            filing: Parsed filing
            output_path: Output Excel file path
            include_errors: Include errors sheet
            
        Raises:
            ImportError: If openpyxl not installed
        """
        if not self.has_openpyxl:
            raise ImportError(
                "openpyxl required for Excel export. "
                "Install with: pip install openpyxl"
            )
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = self.openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Add sheets
        self._create_summary_sheet(wb, filing)
        self._create_facts_sheet(wb, filing)
        self._create_contexts_sheet(wb, filing)
        self._create_units_sheet(wb, filing)
        
        if include_errors and len(filing.errors.errors) > 0:
            self._create_errors_sheet(wb, filing)
        
        # Save workbook
        wb.save(output_path)
        self.logger.info(f"Excel workbook saved: {output_path}")
    
    def _create_summary_sheet(self, wb, filing: ParsedFiling) -> None:
        """Create summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        metadata = filing.metadata
        
        # Title
        ws['A1'] = "XBRL Filing Summary"
        ws['A1'].font = self.openpyxl.styles.Font(size=14, bold=True)
        
        # Filing information
        row = 3
        info = [
            ("Filing ID", metadata.filing_id),
            ("Company", metadata.company_name or "N/A"),
            ("Entity ID", metadata.entity_identifier),
            ("Document Type", metadata.document_type or "N/A"),
            ("Filing Date", str(metadata.filing_date) if metadata.filing_date else "N/A"),
            ("Period End", str(metadata.period_end_date) if metadata.period_end_date else "N/A"),
            ("Market", metadata.market or "N/A"),
            ("", ""),
            ("Total Facts", len(filing.instance.facts)),
            ("Contexts", len(filing.instance.contexts)),
            ("Units", len(filing.instance.units)),
            ("", ""),
            ("Errors", len([e for e in filing.errors.errors if e.severity.value in ['ERROR', 'CRITICAL']])),
            ("Warnings", len([e for e in filing.errors.errors if e.severity.value == 'WARNING'])),
        ]
        
        for label, value in info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if label:
                ws[f'A{row}'].font = self.openpyxl.styles.Font(bold=True)
            row += 1
        
        # set column widths
        ws.column_dimensions['A'].width = self.COLUMN_WIDTH_MEDIUM
        ws.column_dimensions['B'].width = self.COLUMN_WIDTH_EXTRA_WIDE
    
    def _create_facts_sheet(self, wb, filing: ParsedFiling) -> None:
        """Create facts sheet."""
        ws = wb.create_sheet("Facts")
        
        # Extract facts
        facts_data = self.extractor.extract_facts(filing)
        
        if not facts_data:
            ws['A1'] = "No facts found"
            return
        
        # Headers
        headers = list(facts_data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = self.openpyxl.styles.Font(bold=True)
            cell.fill = self.openpyxl.styles.PatternFill(
                start_color="CCCCCC",
                end_color="CCCCCC",
                fill_type="solid"
            )
        
        # Data
        for row, fact_dict in enumerate(facts_data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row, col, fact_dict.get(header))
        
        # Auto-size columns (approximate)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[self.openpyxl.utils.get_column_letter(col)].width = self.COLUMN_WIDTH_NARROW
    
    def _create_contexts_sheet(self, wb, filing: ParsedFiling) -> None:
        """Create contexts sheet."""
        ws = wb.create_sheet("Contexts")
        
        # Extract contexts
        contexts_data = self.extractor.extract_contexts(filing)
        
        if not contexts_data:
            ws['A1'] = "No contexts found"
            return
        
        # Headers
        headers = list(contexts_data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = self.openpyxl.styles.Font(bold=True)
            cell.fill = self.openpyxl.styles.PatternFill(
                start_color="CCCCCC",
                end_color="CCCCCC",
                fill_type="solid"
            )
        
        # Data
        for row, context_dict in enumerate(contexts_data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row, col, context_dict.get(header))
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[self.openpyxl.utils.get_column_letter(col)].width = self.COLUMN_WIDTH_NARROW
    
    def _create_units_sheet(self, wb, filing: ParsedFiling) -> None:
        """Create units sheet."""
        ws = wb.create_sheet("Units")
        
        # Extract units
        units_data = self.extractor.extract_units(filing)
        
        if not units_data:
            ws['A1'] = "No units found"
            return
        
        # Headers
        headers = list(units_data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = self.openpyxl.styles.Font(bold=True)
            cell.fill = self.openpyxl.styles.PatternFill(
                start_color="CCCCCC",
                end_color="CCCCCC",
                fill_type="solid"
            )
        
        # Data
        for row, unit_dict in enumerate(units_data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row, col, unit_dict.get(header))
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[self.openpyxl.utils.get_column_letter(col)].width = self.COLUMN_WIDTH_NARROW
    
    def _create_errors_sheet(self, wb, filing: ParsedFiling) -> None:
        """Create errors sheet."""
        ws = wb.create_sheet("Errors")
        
        # Headers
        ws['A1'] = "Severity"
        ws['B1'] = "Message"
        ws['C1'] = "Location"
        ws['D1'] = "Category"
        
        for cell in [ws['A1'], ws['B1'], ws['C1'], ws['D1']]:
            cell.font = self.openpyxl.styles.Font(bold=True)
            cell.fill = self.openpyxl.styles.PatternFill(
                start_color="CCCCCC",
                end_color="CCCCCC",
                fill_type="solid"
            )
        
        # Data
        row = 2
        for error in filing.errors.errors:
            ws[f'A{row}'] = error.severity.value
            ws[f'B{row}'] = error.message
            ws[f'C{row}'] = error.source_file or ""
            ws[f'D{row}'] = error.category.value
            row += 1
        
        # set column widths
        ws.column_dimensions['A'].width = self.COLUMN_WIDTH_NARROW
        ws.column_dimensions['B'].width = self.COLUMN_WIDTH_VERY_WIDE
        ws.column_dimensions['C'].width = self.COLUMN_WIDTH_WIDE
        ws.column_dimensions['D'].width = self.COLUMN_WIDTH_MEDIUM


__all__ = ['ExcelExporter']