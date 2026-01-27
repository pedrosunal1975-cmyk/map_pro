# Path: output/excel_exporter.py
"""
Excel Exporter

Exports financial statements to Excel format with hierarchical folder structure.
"""

import logging
from pathlib import Path
from typing import Callable

from ..loaders.parser_output import ParsedFiling
from ..mapping.statement.models import Statement, StatementSet
from ..mapping.constants import NetworkCategory


class ExcelExporter:
    """
    Exports statements to Excel format.
    
    Creates folder structure:
    - core_statements/
    - details/
    - other/
    
    Each statement becomes an Excel file with:
    - Formatted headers
    - Proper column widths
    - Data types preserved
    """
    
    def __init__(self):
        """Initialize Excel exporter."""
        self.logger = logging.getLogger('output.excel_exporter')
        
        # Check if openpyxl is available
        try:
            from openpyxl import Workbook
            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
            self.logger.warning("openpyxl not available, Excel export will be skipped")
    
    def export(
        self,
        statement_set: StatementSet,
        parsed_filing: ParsedFiling,
        output_folder: Path,
        filename_creator: Callable
    ) -> list[str]:
        """
        Export statements to Excel with folder structure.
        
        Args:
            statement_set: Set of statements to export
            parsed_filing: Parsed filing data
            output_folder: Base output folder
            filename_creator: Function to create filenames
            
        Returns:
            List of created Excel file paths
        """
        if not self.openpyxl_available:
            self.logger.warning("Skipping Excel export - openpyxl not available")
            return []
        
        excel_paths = []
        
        # Create folder structure
        core_folder = output_folder / 'core_statements'
        details_folder = output_folder / 'details'
        other_folder = output_folder / 'other'
        
        core_folder.mkdir(parents=True, exist_ok=True)
        details_folder.mkdir(parents=True, exist_ok=True)
        other_folder.mkdir(parents=True, exist_ok=True)
        
        # Export each statement to Excel
        for statement in statement_set.statements:
            classification = statement.metadata.get('classification', {})
            category = classification.get('category', NetworkCategory.UNKNOWN)
            
            # Determine folder
            if category == NetworkCategory.CORE_STATEMENT:
                folder = core_folder
            elif category == NetworkCategory.DETAIL:
                folder = details_folder
            else:
                folder = other_folder
            
            # Create filename and export
            filename = filename_creator(statement)
            excel_path = self._export_file(statement, folder, filename)
            excel_paths.append(str(excel_path))
        
        self.logger.info(f"Exported {len(excel_paths)} Excel files")
        return excel_paths
    
    def _export_file(
        self,
        statement: Statement,
        output_folder: Path,
        filename: str
    ) -> Path:
        """
        Export single statement to Excel file.
        
        Args:
            statement: Statement to export
            output_folder: Folder to save to
            filename: Filename (without extension)
            
        Returns:
            Path to created file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        excel_path = output_folder / f"{filename}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"
        
        # Header row with formatting
        headers = [
            'Concept', 'Value', 'Display Value', 'Formatted Value',
            'Context Ref', 'Unit Ref', 'Decimals', 'Scaling Factor',
            'Level', 'Parent Concept', 'Order'
        ]
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_idx, fact in enumerate(statement.facts, 2):
            ws.cell(row=row_idx, column=1, value=fact.concept)
            ws.cell(row=row_idx, column=2, value=fact.value)
            ws.cell(row=row_idx, column=3, value=fact.display_value or '')
            ws.cell(row=row_idx, column=4, value=fact.formatted_value or '')
            ws.cell(row=row_idx, column=5, value=fact.context_ref)
            ws.cell(row=row_idx, column=6, value=fact.unit_ref or '')
            ws.cell(row=row_idx, column=7, value=fact.decimals or '')
            ws.cell(row=row_idx, column=8, value=fact.scaling_factor or '')
            ws.cell(row=row_idx, column=9, value=fact.level)
            ws.cell(row=row_idx, column=10, value=fact.parent_concept or '')
            ws.cell(row=row_idx, column=11, value=fact.order or '')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column].width = adjusted_width
        
        # Save workbook
        wb.save(excel_path)
        return excel_path